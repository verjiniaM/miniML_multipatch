# -*- coding: utf-8 -*-
"""
Created on Tue Feb 15 09:57:32 2022.

@author: imbroscb
"""
import copy
from datetime import datetime
import time
import shutil
from scipy import signal
import os
import warnings
warnings.filterwarnings("ignore")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import pandas as pd
from sklearn.metrics.pairwise import euclidean_distances
from miniML_functions import (get_event_peak, get_event_baseline, get_event_onset, get_event_risetime,
                              get_event_halfdecay_time, get_event_charge)

class ParametersCalculator():
    """
    Calculate PSCs amplitude, rise time and decay time.
    """

    def __init__(self, signal_lp, x_dp, y, fs):
        self.x_dp = x_dp
        self.signal_lp = signal_lp
        self.y = y
        self.fs = fs
        self.amplitude_stdev()
        self.rise_time()
        self.rise_time_avg()
        self.decay_time_avg()

    def amplitude_stdev(self):
        """
        Calculate the amplitude of all PSCs and the stdev of the baseline sig.

        Each event trough is subtracted to a local baseline.
        Returns
        -------: 
            amplitude (np.array): the amplitude of each detected event
            amp_abs (np.array): the absolute amplitude of each detected event
            means (list): mean of each patch of signal without events
            stdev (list): std of each patch of signal without events
        """
        trace_within_events_dp = []
        for ev in self.x_dp:
            delta_left = int(min((self.fs / 1000) * 10, ev))
            delta_right = int(
                min((self.fs / 1000) * 50, self.signal_lp.shape[0] - (ev + 1)))
            delta_left = max(0, delta_left)
            delta_right = max(0, delta_right)
            temp = np.linspace(ev - delta_left, ev + delta_right - 1,
                               delta_left + delta_right).astype('int')
            trace_within_events_dp.extend(temp)

        signal_wo_events = copy.deepcopy(self.signal_lp)
        signal_wo_events[trace_within_events_dp] = np.nan
        signal_wo_events = signal_wo_events.astype('float64')

        # get the amplitude of the detected events
        amplitude = np.zeros((len(self.x_dp)))
        amp_abs = np.zeros((len(self.x_dp)))
        for ev_idx, _ in enumerate(self.x_dp):
            delta_left = int(min(self.fs, self.x_dp[ev_idx]))
            delta_right = int(
                min(self.fs, signal_wo_events.shape[0] - self.x_dp[ev_idx]))
            local_baseline = np.nanmean(signal_wo_events[
                self.x_dp[ev_idx] - delta_left:
                self.x_dp[ev_idx] + delta_right])
            amplitude[ev_idx] = local_baseline - self.y[ev_idx]
            amp_abs[ev_idx] = np.abs(local_baseline - self.y[ev_idx])

        stdev, means = [], []
        # get a stdev value from 100 ms local signal every sec
        for seg in range(signal_wo_events.shape[0] // self.fs):
            delta_left = int(min((self.fs // 1000) * 50, seg * self.fs))
            delta_right = int(min((self.fs / 1000) * 50,
                                  signal_wo_events.shape[0] - seg * self.fs))
            current_stdev = np.nanstd(
                signal_wo_events[seg * self.fs - delta_left:
                                 seg * self.fs + delta_right])
            current_mean = np.nanmean(
                signal_wo_events[seg * self.fs - delta_left:
                                 seg * self.fs + delta_right])
            stdev.append(current_stdev)
            means.append(current_mean)

        return amplitude, amp_abs, means, stdev

    def rise_time_avg(self):
        """
        Calculate the 10-90% rise time.

        To calculate the rise time the average of the detected PSCs are used.
        Cutouts of PSCs (20 ms up to the trough) are used to calculate the
        average signal. PSCs close to the beginning of the recording (<20ms)
        are excluded from calculating the average signal.
        """
        cutouts_rise = []
        x_dp = self.x_dp
        for i in range(len(x_dp)):

            # create cutouts from the rise phase
            steps = int(self.fs / 1000 * 20)
            delta_left = int(min(steps, x_dp[i]))
            ydata = self.signal_lp[x_dp[i] - delta_left:x_dp[i]].reshape(-1,)

            if len(ydata) == steps:
                cutouts_rise.append(ydata)

        cutout_rise_mean = np.mean(np.array(cutouts_rise), axis=0)

        # find the values range
        crm_range = np.max(cutout_rise_mean) - np.min(cutout_rise_mean)

        # normalize the mean trace between 0 and 1 and invert it
        crm_scaled = (cutout_rise_mean - np.min(cutout_rise_mean)) / crm_range
        crm_scaled_inv = 1 - crm_scaled

        # get datapoints number (dp) within 10-90 %
        found_10 = False
        dp_10 = None
        dp_90 = None
        dp_rev_idx = len(crm_scaled_inv) - 1
        for _ in range(len(crm_scaled_inv) - 1):
            if crm_scaled_inv[dp_rev_idx] >= 0.9:
                if crm_scaled_inv[dp_rev_idx - 1] < 0.9:
                    dp_90 = dp_rev_idx
            if crm_scaled_inv[dp_rev_idx] >= 0.1:
                if (crm_scaled_inv[dp_rev_idx - 1] < 0.1) and (not found_10):
                    dp_10 = dp_rev_idx
                    found_10 = True
            dp_rev_idx -= 1

        try:
            rise_time_dp = dp_90 - dp_10

            # calculate the rise time (10-90 %)
            rise_time = rise_time_dp / (self.fs / 1000)

        except TypeError:
            rise_time = np.nan

        return rise_time, dp_10, dp_90

    def rise_time(self):
        """
        Calculate the 10-90% rise time.

        To calculate the rise time the average of the detected PSCs are used.
        Cutouts of PSCs (20 ms up to the trough) are used to calculate the
        average signal. PSCs close to the beginning of the recording (<20ms)
        are excluded from calculating the average signal.
        """
   
        x_dp = self.x_dp
        rise_times = []
        for _, x in enumerate(x_dp):

            # create cutouts from the rise phase
            steps = int(self.fs / 1000 * 20)
            delta_left = int(min(steps, x))
            ydata = self.signal_lp[x - delta_left:x].reshape(-1,)

            if len(ydata) == steps:
                crm_range = np.max(ydata) - np.min(ydata)

                # normalize the mean trace between 0 and 1 and invert it
                crm_scaled = (ydata - np.min(ydata)) / crm_range
                crm_scaled_inv = 1 - crm_scaled

                # get datapoints number (dp) within 10-90 %
                found_10 = False
                dp_10 = None
                dp_90 = None
                dp_rev_idx = len(crm_scaled_inv) - 1
                for _ in range(len(crm_scaled_inv) - 1):
                    if crm_scaled_inv[dp_rev_idx] >= 0.9:
                        if crm_scaled_inv[dp_rev_idx - 1] < 0.9:
                            dp_90 = dp_rev_idx
                    if crm_scaled_inv[dp_rev_idx] >= 0.1:
                        if (crm_scaled_inv[dp_rev_idx - 1] < 0.1) and (not found_10):
                            dp_10 = dp_rev_idx
                            found_10 = True
                    dp_rev_idx -= 1
                try:
                    rise_time_dp = dp_90 - dp_10
                    # calculate the rise time (10-90 %)
                    rise_time = rise_time_dp / (self.fs / 1000)
                except TypeError:
                    rise_time = np.nan
            else:
                rise_time = np.nan

            rise_times.append(rise_time)
        return rise_times

    def decay_time_avg(self):
        """
        Calculate the 90-10% decay time.

        To calculate the decay time the average of the detected PSCs are used.
        Cutouts of PSCs (50 ms starting from the trough) are used to calculate
        the average signal. PSCs close to the end of the recording (<50ms) are
        excluded from calculating the average signal.
        """
        cutouts_decay = []
        x_dp = self.x_dp
        for i in range(len(x_dp)):

            # create cutouts from the decay phase
            steps = int(self.fs / 1000 * 50)
            delta_right = int(min(steps, self.signal_lp.shape[0] - (x_dp[i])))
            ydata = self.signal_lp[x_dp[i]:x_dp[i] + delta_right].reshape(-1,)

            if len(ydata) == steps:
                cutouts_decay.append(ydata)

        cutout_decay_mean = np.mean(np.array(cutouts_decay), axis=0)

        # find the values range
        cdm_range = np.max(cutout_decay_mean) - np.min(cutout_decay_mean)

        # normalize the mean trace between 0 and 1
        cdm_scaled = (cutout_decay_mean - np.min(cutout_decay_mean)
                      ) / cdm_range

        # get datapoints number (dp) within 10-90 %
        found_90 = False
        dp_10 = None
        dp_90 = None
        dp_idx = 0
        for _ in range(len(cdm_scaled) - 1):
            if (cdm_scaled[dp_idx] < 0.1):
                if cdm_scaled[dp_idx + 1] >= 0.1:
                    dp_10 = dp_idx + 1
            if cdm_scaled[dp_idx] < 0.9:
                if (cdm_scaled[dp_idx + 1] >= 0.9) and (not found_90):
                    dp_90 = dp_idx + 1
                    found_90 = True
            dp_idx += 1

        try:
            decay_time_dp = dp_90 - dp_10

            # calculate the rise time (10-90 %)
            decay_time = decay_time_dp / (self.fs / 1000)

        except TypeError:
            decay_time = np.nan

        return decay_time, dp_10, dp_90

def init_dicts(attr_names, shape, dtype):
    ''' initialize multiple 1d ndarrays with given shape containing NaNs '''
    empty_d = {}
    for label in attr_names:
        value = -1 if 'int' in str(dtype) else np.NaN
        empty_d[label] = np.full(int(shape), value, dtype=dtype)
    return empty_d

def hann_filter(data, filter_size):
    win = signal.windows.hann(filter_size)
    return signal.convolve(data, win, mode = 'same') / sum(win)

def get_event_properties(data_def, event_locations, sampling,
            window_size = 750, convolve_win = 20, filter: bool = True):
    '''
    Find more detailed event location properties required for analysis. Namely, baseline, event onset,
    peak half-decay and 10 & 90% rise positions. Also extracts the actual event properties, such as
    amplitude or half-decay time.
    
    Parameters
    ------
    filter: bool
        If true, properties are extracted from the filtered data.
    '''
    ### Prepare data
    diffs = np.diff(event_locations, append = len(data_def)) # Difference in points between the event locations
    add_points = int(window_size/3)
    after = window_size + add_points
    positions = event_locations
    
    ### Set parameters for charge calculation
    factor_charge = 4
    num_combined_charge_events = 1
    calculate_charge = False # will be set to True in the loop if double event criteria are fulfilled; not a flag for charge

    if np.any(positions - add_points < 0) or np.any(positions + after >=  len(data_def)):
        raise ValueError('Cannot extract time windows exceeding input data size.')
    
    if filter:
        mini_trace = hann_filter(data = data_def, filter_size = convolve_win)
    else:
        mini_trace = data_def

    mini_trace *= -1 # because the event_direction is negative

    temp_dict = init_dicts(['event_peak_locations', 'event_start', 'min_positions_rise', 'max_positions_rise'], positions.shape[0], dtype = np.int64)
    params_dict = init_dicts(['event_peak_values', 'event_bsls', 'decaytimes', 'charges', 'risetimes', 'half_decay'], positions.shape[0], dtype = np.float64)

    for ix, position in enumerate(positions):
        indices = position + np.arange(-add_points, after)
        data = mini_trace[indices]
            
        if filter:
            data_unfiltered = data_def * -1 # negative event direction
        else:
            data_unfiltered = data
        
        event_peak = get_event_peak(data = data,event_num = ix,add_points = add_points,window_size = window_size,diffs = diffs)
        
        temp_dict['event_peak_locations'][ix] = int(event_peak)
        params_dict['event_peak_values'][ix] = data[event_peak]
        peak_spacer = int(window_size/100)
        params_dict['event_peak_values'][ix] = np.mean(data_unfiltered[int(event_peak-peak_spacer):int(event_peak+peak_spacer)])
        
        baseline, baseline_var = get_event_baseline(data = data, event_num = ix, diffs = diffs, add_points = add_points,
                                                    peak_positions = temp_dict['event_peak_locations'], positions = positions)
        params_dict['event_bsls'][ix] = baseline
        
        onset_position = get_event_onset(data = data, peak_position = event_peak, baseline = baseline,
                                            baseline_var = baseline_var)
        temp_dict['event_start'][ix] = onset_position
        
        risetime, min_position_rise, max_position_rise = get_event_risetime(data = data, peak_position = event_peak,
                                                                            onset_position = onset_position)
        params_dict['risetimes'][ix] = risetime
        temp_dict['min_positions_rise'][ix] = min_position_rise
        temp_dict['max_positions_rise'][ix] = max_position_rise

        level = baseline + (data[event_peak] - baseline) / 2
        if diffs[ix] < add_points: # next event close; check if we can get halfdecay
            right_lim = diffs[ix]+add_points # Right limit is the onset of the next event
            test_arr =  data[event_peak:right_lim]
            if test_arr[test_arr<level].shape[0]: # means that event goes below 50% ampliude before max rise of the next event; 1/2 decay can be calculated
                halfdecay_position, halfdecay_time = get_event_halfdecay_time(data = data[0:right_lim],
                                                                                peak_position = event_peak, baseline = baseline)
            else:
                halfdecay_position, halfdecay_time = np.nan, np.nan
        else:  
            halfdecay_position, halfdecay_time = get_event_halfdecay_time(data = data, peak_position = event_peak, 
                                                                            baseline = baseline)

        params_dict['half_decay'][ix] = halfdecay_position
        params_dict['decaytimes'][ix] = halfdecay_time
        
        # calculate params_dict['charges']
        ### For charge; multiple event check done outside function.
        if ix < positions.shape[0]-1:
            if num_combined_charge_events ==  1: # define onset position for charge calculation
                onset_in_trace = positions[ix] - (add_points-temp_dict['event_start'][ix])
                baseline_for_charge = params_dict['event_bsls'][ix]

            if np.isnan(params_dict['half_decay'][ix]):
                num_combined_charge_events += 1
            
            else:
                ### Get distance from peak to next event location.
                peak_in_trace = positions[ix] + (temp_dict['event_peak_locations'][ix] - add_points)
                next_event_location = positions[ix+1]
                delta_peak_location = next_event_location - peak_in_trace
                
                # determine end of area calculation based on event decay
                endpoint = int(temp_dict['event_peak_locations'][ix] + \
                                factor_charge*(int(params_dict['half_decay'][ix]) - temp_dict['event_peak_locations'][ix]))
                delta_peak_endpoint = endpoint-temp_dict['event_peak_locations'][ix]

                if delta_peak_location > delta_peak_endpoint: # Next event_location further away than the charge window; calculate charge
                    calculate_charge = True
                else:
                    num_combined_charge_events += 1

            if calculate_charge:
                endpoint_in_trace = positions[ix] + (temp_dict['event_peak_locations'][ix] - add_points) + delta_peak_endpoint
                charge = get_event_charge(data = mini_trace, start_point = onset_in_trace, end_point = endpoint_in_trace,
                                           baseline = baseline_for_charge, sampling = sampling)
                
        else: # Handle the last event
            if num_combined_charge_events ==  1: # define onset position for charge calculation
                onset_in_trace = positions[ix] - (add_points - temp_dict['event_start'][ix])
                baseline_for_charge = params_dict['event_bsls'][ix]
            
            peak_in_trace = positions[ix] + (temp_dict['event_peak_locations'][ix] - add_points)
            endpoint = int(temp_dict['event_peak_locations'][ix] + factor_charge*(int(params_dict['half_decay'][ix]) - temp_dict['event_peak_locations'][ix]))
            delta_peak_endpoint = endpoint-temp_dict['event_peak_locations'][ix]
            endpoint_in_trace = positions[ix] + (temp_dict['event_peak_locations'][ix] - add_points) + delta_peak_endpoint
            
            if endpoint_in_trace > mini_trace.shape[0]:
                endpoint_in_trace = mini_trace.shape[0]

            charge = get_event_charge(data = mini_trace, start_point = onset_in_trace, end_point = endpoint_in_trace, baseline = baseline_for_charge, sampling = sampling)
            calculate_charge = True
        if calculate_charge: # Charge was caclulated; check how many potentially overlapping events contributed.
            charge = [charge/num_combined_charge_events]*num_combined_charge_events
            for ix_adjuster in range(len(charge)):
                params_dict['charges'][ix-ix_adjuster] = charge[ix_adjuster]
                        
            # Reset values after calculation
            calculate_charge = False
            num_combined_charge_events = 1

    ## Convert units
    params_dict['event_peak_values'] *=  -1
    params_dict['event_bsls'] *=  -1
    params_dict['risetimes'] *=  sampling
    params_dict['decaytimes'] *=  sampling
    params_dict['charges'] *=  -1

    ## map indices back to original         
    for ix, position in enumerate(positions):
        temp_dict['event_peak_locations'][ix] = int(temp_dict['event_peak_locations'][ix] + event_locations[ix] - add_points)
        temp_dict['event_start'][ix] = int(temp_dict['event_start'][ix] + event_locations[ix] - add_points)
        temp_dict['min_positions_rise'][ix] = int(min_position_rise + event_locations[ix] - add_points)
        temp_dict['max_positions_rise'][ix] = int(max_position_rise + event_locations[ix] - add_points)            
        
        if not np.isnan(params_dict['half_decay'][ix]):
            params_dict['half_decay'][ix] = int(params_dict['half_decay'][ix] + event_locations[ix] - add_points)
    return temp_dict, params_dict

class MiniSpontProof:
    """Give the possibility to check and revise the results."""

    def __init__(self, sum_df_path, num_swps, individ_df_path, time_,
                sampling, signal_lp, original_values, fs):
        '''
        singal_lp is just the name of the varialbe. no low pass filtering performed
        passing raw data and then Hann window filtering it
        '''
        self.sum_df = sum_df_path
        self.individ_df_path = individ_df_path
        self.time = time_
        self.signal_lp = signal_lp
        # self.signal_std_plot = std_aprox
        self.original_values = original_values
        self.num_swps = num_swps
        self.fs = fs
        self.sampling = sampling
        self.fn_prefix = individ_df_path[individ_df_path.rfind('/')+1:individ_df_path.rfind('_individual')-2]
        self.chan = individ_df_path[individ_df_path.rfind('_individual')-1:individ_df_path.rfind('_individual')]
        self.model_path = '/Users/verjim/miniML_multipatch/models/transfer_learning/human_pyramids_L2_3/2024_lstm_transfer_dec_2024.h5'
        self.last_save_time = time.time()  # Track last save time
        self.auto_save_interval = 30  # Auto-save every 30 seconds
        self.init_window()

    def init_window(self):
        """Plot signal_lp and detected m/s PSCs."""
        self.fig = plt.figure(figsize=(16, 8))
        self.ax = self.fig.add_subplot(1, 1, 1)
        data_plot = hann_filter(self.signal_lp, filter_size = 20)
        self.ax.plot(self.time * self.fs/1000 , data_plot, c = 'black')
        # self.ax.plot(self.time[:-900], self.signal_std_plot[900:],
        #              c = 'red', linestyle = '--', label = 'std aprox')
        x_lines = np.linspace(self.time[-1] * self.fs/1000/self.num_swps,
                              self.time[-1] * self.fs/1000, self.num_swps)
        
        self.ax.vlines(x = x_lines, ymin = min(self.signal_lp),
                       ymax = max(self.signal_lp), colors = 'orange',
                       linestyles='-', label = 'sweep ends')
        if len(self.original_values) > 0:
            self.ax.scatter(self.original_values[:, 0],
                            self.original_values[:, 1], c='orange', s=70)
        self.fig.suptitle(self.fn_prefix + ' ch:' + self.chan)
        self.ax.legend(loc='upper left')
        self.ax.title.set_text('Press "u" to update the results.xlsx file.')
        self.ax.set_xlabel('Time (ms)')
        self.ax.set_xlabel('datapoints')
        self.ax.set_ylabel('Signal (pA)')

        # the following lines are necessary for the manual correction
        self.updated_values = self.original_values
        self.false_positives = []
        # next line specifies that when a key is pressed, the plot window is active
        # and when active, the plot calls the create_new_dataset method
        plt.connect('key_press_event', self.create_new_dataset)

        # Set the zoom/limits before showing
        self.ax.set_xlim(-25*20, 550 * 20) # X-axis limits
        self.ax.set_ylim(np.mean(self.signal_lp) - 35, np.mean(self.signal_lp) + 5)
        plt.show()

    def create_new_dataset(self, event):
        """
        Add points, remove points and update the results.xlsx file.

        The keyboard button 'a' is used to add false negatives.
        The keyboard button 'd' is used to remove false positives.
        The keyboard button 'u' is used to update the results.xlsx file.
        """
        # add a new event in response to the keyboard button 'a'
              # Check if 30 seconds have passed since last save
        current_time = time.time()
        if current_time - self.last_save_time >= self.auto_save_interval:
            print("Auto-saving results (30 seconds elapsed)...")
            self._update_results()
            self.last_save_time = current_time
        if event.key == 'a':
            new_xy_event = [event.xdata, event.ydata]
            # define the current events
            xy_events_tp = self.updated_values

            # update new values: current events + new event
            if len(self.updated_values) > 0:
                self.updated_values = np.insert(xy_events_tp, 0, new_xy_event,
                                                axis=0)
            else:
                self.updated_values = np.array(new_xy_event).reshape(1, -1)

            # plot the new event
            self.ax.scatter(event.xdata, event.ydata, c='blue', s=70)
            self.fig.canvas.draw()

        # delete an event in response to the keyboard button 'd',
        # if a data point is close enough to the clicking position
        if event.key == 'd':

            xy_click = np.array([event.xdata, event.ydata])
            xy_events = self.updated_values

            # look for event detection (data point) closest to the click
            closest_point = xy_events[0]
            min_euc = 999999
            idx_to_delete = 0
            for idx in range(len(xy_events)):
                euc_distance = euclidean_distances(
                    xy_events[idx].reshape(1, -1), xy_click.reshape(1, -1))
                if euc_distance < min_euc:
                    min_euc = euc_distance
                    closest_point = xy_events[idx]
                    idx_to_delete = idx

            # if the closest data point is close enough, it will be deleted
            if min_euc < 50:
                self.updated_values = np.delete(xy_events, idx_to_delete,
                                                axis=0)
                # plot the the deleted data point in red
                self.ax.scatter(closest_point[0], closest_point[1], c='red', s=70)
                self.fig.canvas.draw()

                # define the current false positives data point
                xy_events_fp = self.false_positives

                # update false positives
                if len(self.false_positives) > 0:
                    self.false_positives = np.insert(xy_events_fp, 0,
                                                     closest_point, axis=0)
                else:
                    self.false_positives = np.array(
                        closest_point).reshape(1, -1)

        if event.key == 'u':
            self._update_results()
            self.last_save_time = time.time()  # Update last save time

    def _update_results(self):
        # get the updated x, y values
        # if 'a' not pressed before, updated = org
        x = self.updated_values[:, 0].astype('float64')
        y = self.updated_values[:, 1].astype('float64')

        # sort the updated x, y values (for ascending x)
        x, y = zip(*sorted(zip(x, y)))
        x = np.array(x)
        y = np.array(y)

        # get the datapoint from the ms
        x_dp = (x).astype('int')

        # calculate the interevent interval
        x_pd = pd.Series(x)
        interevent_interval = (x_pd.shift(-1) - x_pd) / (self.fs / 1000)
        average_interevent_interval = np.nanmean(interevent_interval)

        # instantiate parameters
        params_dict = get_event_properties(self.signal_lp, x_dp, self.sampling)[1]

        amp_miniML = params_dict['event_peak_values'] - params_dict['event_bsls']
        abs_amp_miniML = np.abs(params_dict['event_peak_values'] - params_dict['event_bsls'])
        # set amplitude to nan, when events too close to each other
        # amplitude[interevent_interval < 45] = np.nan

        # from miniML - wrong, needs evaluation
        # rise_times = params_dict['risetimes']
        # decay_times = params_dict['decaytimes']

        # from barb
        params_barb_raw = ParametersCalculator(self.signal_lp, x_dp, y, self.fs)

        hann_signal = hann_filter(self.signal_lp, filter_size = 20)
        params_barb_hann = ParametersCalculator(hann_signal, x_dp, y, self.fs)
        amp_barb_hann, abs_amp_barb_hann, mean_hann, stdev_hann = params_barb_hann.amplitude_stdev()

        sweep_ends = np.linspace(self.time[-1] * self.fs/1000/self.num_swps,
                        self.time[-1] * self.fs/1000, self.num_swps)
        from_swp = np.searchsorted(a = sweep_ends, v = x, side = 'right') # a[i-1] <= v < a[i]

        ### INDIVIDUAL DF
        wb_i = load_workbook(self.individ_df_path)
        ws_revised = wb_i['Sheet1'] # create sheet

        headers = ['x (dp)', 'y (pA)', 'Interevent interval (ms)',
                    'amp_miniML_pA', 'abs_amp_mimiML_pA', 
                    'amp_barb_pA', 'abs_amp_barb_pA', 'sweep_num']

        for col, header in enumerate(headers, 1):
            ws_revised.cell(1, col).value = header
        params_to_add = [x, y, interevent_interval, 
                         amp_miniML, abs_amp_miniML,
                        amp_barb_hann, abs_amp_barb_hann,
                        from_swp]
        # fill values
        for i, param in enumerate(params_to_add):
            for row in range(len(param)):
                ws_revised.cell(row + 2, i+1).value = param[row]
    
        wb_i.save(self.individ_df_path)
        wb_i.close()
        
        # SUMMARY DF
        wb_summary = load_workbook(self.sum_df)
        ws_summary = wb_summary['Summary results']
                    
        for col in range(ws_summary.max_column):
            if ws_summary.cell(1, col + 1).value == 'filename':
                col_rec_file = col + 1
            if ws_summary.cell(1, col + 1).value == 'channel':
                col_channel = col + 1
            if ws_summary.cell(1, col + 1).value == 'avg_IEI_ms':
                col_inter_ev_inter = col + 1
            if ws_summary.cell(1, col + 1).value == 'avg_amp_pA':
                col_amplitude = col + 1
            if ws_summary.cell(1, col + 1).value == 'abs_avg_amp_pA':
                col_abs_amplitude = col + 1
            if ws_summary.cell(1,col + 1).value == 'avg_rise_times_10_90_ms_barb':
                col_rise_time = col + 1
            if ws_summary.cell(1, col + 1).value == 'avg_decay_times_10_90_ms_barb':
                col_decay_time = col + 1
            if ws_summary.cell(1, col+1).value == 'raw_mean_bl_signal_pA':
                col_avg_bl = col + 1
            if ws_summary.cell(1, col+1).value == 'hann_mean_bl_signal_pA':
                col_avg_hann = col + 1
            if ws_summary.cell(1, col + 1).value == 'raw_stdv_bl_signal_pA':
                col_std_raw= col + 1
            if ws_summary.cell(1, col+1).value == 'hann_stdv_bl_signal_pA':
                col_std_hann = col + 1
            if ws_summary.cell(1, col + 1).value == 'man_revised':
                col_man_rev = col + 1
            if ws_summary.cell(1, col + 1).value == 'time_analysis_end':
                col_date = col + 1


        for row in range(ws_summary.max_row):
            cell_value = ws_summary.cell(row + 1, col_rec_file).value
            # find corresponding fn and chan
            if cell_value[:-4] == self.fn_prefix:
                channel_value = int(ws_summary.cell(row + 1, col_channel).value)
                if channel_value == int(self.chan):
                    row_to_modify = row + 1
                    ws_summary.cell(row_to_modify, col_inter_ev_inter).value = average_interevent_interval
                    ws_summary.cell(row_to_modify, col_amplitude).value = np.nanmean(amp_barb_hann)
                    ws_summary.cell(row_to_modify, col_abs_amplitude).value = np.nanmean(abs_amp_barb_hann)
                    ws_summary.cell(row_to_modify, col_rise_time).value =  params_barb_hann.rise_time_avg()[0]
                    ws_summary.cell(row_to_modify, col_decay_time).value = params_barb_hann.decay_time_avg()[0]
                    ws_summary.cell(row_to_modify, col_avg_bl).value = np.nanmean(params_barb_raw.amplitude_stdev()[2])
                    ws_summary.cell(row_to_modify, col_avg_hann).value = np.nanmean(mean_hann)
                    ws_summary.cell(row_to_modify, col_std_raw).value = np.nanmean(params_barb_raw.amplitude_stdev()[3])
                    ws_summary.cell(row_to_modify, col_std_hann).value = np.nanmean(stdev_hann)
                    ws_summary.cell(row_to_modify, col_man_rev).value = 'yes'
                    ws_summary.cell(row_to_modify, col_date).value = datetime.now().strftime("%d.%m.%y %H:%M ")
                    break

        # save and close wb
        wb_summary.save(self.sum_df)
        print('The sum_results.xlsx file has been updated')
        wb_summary.close()
        self.last_save_time = time.time()

def lowpass(data, cutoff_frequency, fs=20000, order=2):
    """
    Low pass filters the input signal.

    inputs:
        data = the signal to be filtered. It needs to have shape [-1, 1]
        cutoff_frequency = the cutoff frequency
        fs = the sampling rate (Hz)
        order = polynomial
    output:
        the filtered_data
    """
    nyq = fs * 0.5
    b, a = signal.butter(order, cutoff_frequency / nyq)
    if len(data.shape) > 1:
        if data.shape[0] > data.shape[1]:
            ref_filt = signal.lfilter(b, a, data, axis=0)
        else:
            ref_filt = signal.lfilter(b, a, data, axis=1)
    else:
        ref_filt = signal.lfilter(b, a, data, axis=0)
    return ref_filt


def copy_event_files(meta_df, dest_dir,
                    human_dir = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/'):
    '''
    copies the files to recordings dir
    '''
    counter = 0
    for i, OP in enumerate(meta_df.OP):
        patcher =  meta_df['patcher'][i]

        if patcher == 'Verji': 
            work_dir = human_dir + 'data_verji/' + OP + '/'
        elif patcher == 'Rosie':
            work_dir = human_dir + 'data_rosie/' + OP + '/'

        file_path = os.path.join(work_dir, meta_df['Name of recording'][i])
        if not os.path.exists(dest_dir + meta_df['Name of recording'][i]):
            shutil.copy(file_path, dest_dir)
            counter +=1
        else:
            continue
    print(f'copied {counter} files to {file_path}')

